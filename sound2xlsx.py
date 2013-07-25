import json
import urllib
from openpyxl import load_workbook
import soundcloud
import datetime

def auth():
        client = soundcloud.Client(
                client_id='',
                client_secret='',
                username='',
                password='')
        test = client.get('/me/activities/tracks/').collection
        return test

trks = []
act = auth()
try:
	for track in act:
        	test4 = track['origin']['permalink_url']
		trks.append(test4)
		print test4
except:
	pass

#today = datetime.date.today()
#date = today.strftime('%m-%d-%y')
#users = []
#favn = []
#list = []
y = 2
y2 = 2
tel = {}
uri = 'http://api.soundcloud.com/users/dasu/followings.json?client_id=YOUR_CLIENT_ID'
bytes = urllib.urlopen(uri)
test = bytes.read()
m = json.loads(test)

for name in m:
  	user = name['permalink']
	fav = name['public_favorites_count']
#	users.append(user)
#	favn.append(fav)
	tel[user] = fav
	
#if users == []:
#	print 'Nothing found'
#else:
#	print (",".join(users))
#	print (",".join(str(w) for w in favn))

wb2 = load_workbook('test1.xlsx')
sheet1 = wb2.get_sheet_by_name(name= 'Sheet1')
sheet2 = wb2.get_sheet_by_name(name= 'Sheet2')

x = sheet1.get_highest_column()
x2 = sheet2.get_highest_column()

#print 'x = ' + str(x)
#print 'y = ' + str(y)
for row in sheet1.range('A3:A25'):
	for p in row:
		sheet1.cell(row = y, column = x).value = tel[p.value]
		#cv.value = tel[p.value]
#		print cv.value
		y += 1		
#		print y
#		list.append(p.value)
#		print tel[p.value]
		#print p.value + ' has ' + str(tel[p.value])
for fdsg in trks:		
	sheet2.cell(row = y2, column = x2).value = fdsg
	y2 += 1
#print trks

#datev = sheet1.cell(row = 2, column = x)
#datev.value = date
sheet1.cell(row = 1, column = x).value = datetime.datetime.today()
#datev2 = sheet2.cell(row = 2, column = x2)
#datev2.value = date
sheet2.cell(row = 1, column = x2).value = datetime.datetime.today()

#for hurr in sheet1.range('A1:M10'):
#	for durr in hurr:
#		print durr.value

#while y < len(list):
#	if list[y] == users[0]:
#		print list[y] + ' true'
#		y += 1
#	else:
#		print list[y] + ' false'
#		y += 1
#		
#if not sheet1.cell('A3').value:
#	print 'Cell has no value'
#else:
#	print 'Cell has value:'
#	print sheet1.cell('A3').value

#print tel['posse-1']
#print sheet1.get_highest_column()
wb2.save('test1.xlsx')

